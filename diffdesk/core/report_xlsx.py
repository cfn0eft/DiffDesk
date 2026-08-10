"""色付きExcel差分レポートの生成(openpyxl)。

大規模データ対応のため write_only モードで生成する(セルオブジェクトを
メモリに保持しないため、10万行規模でもメモリ・時間ともに実用的)。
列幅は先頭数百行のサンプルから決める。
"""
from __future__ import annotations

import io as _io

from .model import DiffResult
from .upsert import build_report_table

_FILL_ONLY_A = "C6EFCE"   # 緑: Aのみ(insert候補)
_FILL_ONLY_B = "FFC7CE"   # 赤: Bのみ
_FILL_CHANGED = "FFEB9C"  # 黄: 変更セル
_FILL_HEADER = "D9E1F2"

_WIDTH_SAMPLE_ROWS = 300


def _col_widths(columns: list[str], rows: list[list[str]]) -> list[float]:
    """先頭サンプル行から列幅を決める(全行走査はしない)。"""
    widths = [len(str(c)) for c in columns]
    for row in rows[:_WIDTH_SAMPLE_ROWS]:
        for i, cell in enumerate(row):
            n = len(str(cell))
            if n > widths[i]:
                widths[i] = n
    return [min(w * 1.6 + 2, 40) for w in widths]


def build_xlsx_report(diff: DiffResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook(write_only=True)
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=_FILL_HEADER)
    fill_a = PatternFill("solid", fgColor=_FILL_ONLY_A)
    fill_b = PatternFill("solid", fgColor=_FILL_ONLY_B)
    fill_ch = PatternFill("solid", fgColor=_FILL_CHANGED)

    # --- サマリーシート
    ws = wb.create_sheet("サマリー")
    s = diff.summary
    head = [WriteOnlyCell(ws, v) for v in ("項目", "件数")]
    for c in head:
        c.font = bold
    ws.append(head)
    for label, key in [("Aのみ(insert候補)", "only_a"), ("Bのみ", "only_b"),
                       ("変更", "changed"), ("一致", "same"),
                       ("キー重複(A)", "duplicates_a"), ("キー重複(B)", "duplicates_b"),
                       ("キー空(A)", "empty_key_a"), ("キー空(B)", "empty_key_b")]:
        ws.append([label, s[key]])
    ws.append([])
    ws.append(["ファイルA", diff.name_a])
    ws.append(["ファイルB", diff.name_b])

    # --- 詳細シート
    report = build_report_table(diff)
    ws2 = wb.create_sheet("差分詳細")
    for i, w in enumerate(_col_widths(report.columns, report.rows), start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    head = []
    for v in report.columns:
        c = WriteOnlyCell(ws2, v)
        c.font = bold
        c.fill = header_fill
        head.append(c)
    ws2.append(head)

    n_keys = len(diff.mapping.key_pairs)
    value_pairs = diff.mapping.value_pairs

    for rd, report_row in zip(diff.rows, report.rows):
        status = report_row[0]
        if status == "Aのみ":
            row = [WriteOnlyCell(ws2, v) for v in report_row]
            for c in row:
                c.fill = fill_a
        elif status == "Bのみ":
            row = [WriteOnlyCell(ws2, v) for v in report_row]
            for c in row:
                c.fill = fill_b
        elif status == "変更":
            row = [WriteOnlyCell(ws2, v) for v in report_row]
            row[0].fill = fill_ch
            changed_cols = {cd.col_a for cd in rd.cell_diffs}
            for vi, p in enumerate(value_pairs):
                if p.col_a in changed_cols:
                    base = n_keys + 1 + vi * 2  # 0始まり: 状態+キー列の後のA/Bペア
                    row[base].fill = fill_ch
                    row[base + 1].fill = fill_ch
        else:
            row = report_row  # 一致行はスタイルなし(そのままappendが最速)
        ws2.append(row)

    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()
