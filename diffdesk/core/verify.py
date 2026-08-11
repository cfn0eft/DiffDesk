"""投入検証: Data Loader投入後のレコード件数照合と差異判定。

前提: A = 投入元(正マスタ)、B = 投入後にSalesforceからエクスポートしたデータ。
DiffResult をもとに「投入が成功したか」を機械判定する。

- 未投入      = only_a  (Aにあるのに B に存在しない)
- 値差異      = changed (キーは一致するが値が異なる)
- 想定外      = only_b  (Bにだけ存在する。既存レコードが混ざる運用では許容可)
- 照合不能    = キー重複・キー空(突き合わせから除外されたもの)
"""
from __future__ import annotations

import io as _io
from dataclasses import dataclass, field

from .model import DiffResult, Table
from .upsert import build_report_table


@dataclass
class VerificationResult:
    passed: bool
    rows_a: int              # 突き合わせ対象のA件数(重複・空キー除外後)
    rows_b: int
    matched: int             # キーが一致した件数
    same: int
    changed: int
    only_a: int              # 未投入(既知除く)
    only_b: int              # 想定外(Bのみ、既知除く)
    unmatchable_a: int       # A側のキー重複+キー空
    unmatchable_b: int
    only_b_is_error: bool
    known_rows: int = 0      # 既知として容認された欠落レコード数
    known_cells: int = 0     # 既知として容認されたセル差分数
    problems: list[str] = field(default_factory=list)  # 人向けの指摘文

    def to_dict(self) -> dict:
        return {
            "passed": self.passed,
            "rows_a": self.rows_a, "rows_b": self.rows_b,
            "matched": self.matched, "same": self.same, "changed": self.changed,
            "only_a": self.only_a, "only_b": self.only_b,
            "unmatchable_a": self.unmatchable_a, "unmatchable_b": self.unmatchable_b,
            "only_b_is_error": self.only_b_is_error,
            "known_rows": self.known_rows, "known_cells": self.known_cells,
            "problems": self.problems,
            "match_rate": (self.same / self.rows_a) if self.rows_a else 0.0,
        }


def build_verification(diff: DiffResult, *, only_b_is_error: bool = True) -> VerificationResult:
    s = diff.summary
    only_a = only_b = changed = same = 0
    known_rows = known_cells = 0
    for rd in diff.rows:
        known_cells += len(rd.known_diffs)
        if rd.status == "only_a":
            if rd.known:
                known_rows += 1
            else:
                only_a += 1
        elif rd.status == "only_b":
            if rd.known:
                known_rows += 1
            else:
                only_b += 1
        elif rd.status == "changed":
            changed += 1
        else:
            same += 1
    matched = same + changed
    rows_a = s["only_a"] + matched  # 既知含む物理件数
    rows_b = s["only_b"] + matched
    unmatchable_a = s["duplicates_a"] + s["empty_key_a"]
    unmatchable_b = s["duplicates_b"] + s["empty_key_b"]

    problems: list[str] = []
    if only_a:
        problems.append(f"未投入(Aのみ)が {only_a} 件あります。投入漏れの可能性があります。")
    if changed:
        problems.append(f"値差異が {changed} 件あります。投入時の変換・上書きを確認してください。")
    if only_b:
        msg = f"B(Salesforce)のみのレコードが {only_b} 件あります。"
        if only_b_is_error:
            problems.append(msg + "新規オルグへの投入であれば想定外です。")
        else:
            problems.append("(許容) " + msg + "既存レコードとして許容されています。")
    if known_rows or known_cells:
        detail = []
        if known_rows:
            detail.append(f"欠落{known_rows}件")
        if known_cells:
            detail.append(f"セル差分{known_cells}箇所")
        problems.append(f"(既知) 登録済みの既知差分({'、'.join(detail)})は問題に含めていません。")
    if unmatchable_a:
        problems.append(f"A側にキー重複・キー空が {unmatchable_a} 件あり、照合できていません。")
    if unmatchable_b:
        problems.append(f"B側にキー重複・キー空が {unmatchable_b} 件あり、照合できていません。")

    passed = (only_a == 0 and changed == 0 and unmatchable_a == 0 and unmatchable_b == 0
              and (only_b == 0 or not only_b_is_error))
    return VerificationResult(
        passed=passed, rows_a=rows_a, rows_b=rows_b, matched=matched,
        same=same, changed=changed, only_a=only_a, only_b=only_b,
        unmatchable_a=unmatchable_a, unmatchable_b=unmatchable_b,
        only_b_is_error=only_b_is_error,
        known_rows=known_rows, known_cells=known_cells, problems=problems,
    )


def _problem_statuses(only_b_is_error: bool) -> set[str]:
    statuses = {"only_a", "changed"}
    if only_b_is_error:
        statuses.add("only_b")
    return statuses


def build_verification_table(diff: DiffResult, *, only_b_is_error: bool = True) -> Table:
    """問題レコードのみの検証レポートTable(全行レポートの絞り込み版)。"""
    report = build_report_table(diff)
    status_ja = {"only_a": "Aのみ", "only_b": "Bのみ", "changed": "変更"}
    wanted = {status_ja[s] for s in _problem_statuses(only_b_is_error)}
    rows = [row for row in report.rows if row[0] in wanted]
    return Table(columns=report.columns, rows=rows, name="投入検証レポート")


def build_verification_xlsx(diff: DiffResult, *, only_b_is_error: bool = True) -> bytes:
    """検証サマリー+問題行のみの色付きxlsxレポート。"""
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    from .report_xlsx import _col_widths

    v = build_verification(diff, only_b_is_error=only_b_is_error)
    wb = Workbook(write_only=True)
    bold = Font(bold=True)
    big = Font(bold=True, size=16)
    fill_ok = PatternFill("solid", fgColor="C6EFCE")
    fill_ng = PatternFill("solid", fgColor="FFC7CE")
    fill_header = PatternFill("solid", fgColor="D9E1F2")
    fills = {
        "Aのみ": PatternFill("solid", fgColor="C6EFCE"),
        "Bのみ": PatternFill("solid", fgColor="FFC7CE"),
        "変更": PatternFill("solid", fgColor="FFEB9C"),
    }

    ws = wb.create_sheet("検証結果")
    verdict = WriteOnlyCell(ws, "✔ 投入OK(全件一致)" if v.passed else "✖ 要確認(差異あり)")
    verdict.font = big
    verdict.fill = fill_ok if v.passed else fill_ng
    ws.append([verdict])
    ws.append([])
    for label, value in [
        ("投入元(A) 照合対象件数", v.rows_a),
        ("Salesforce(B) 照合対象件数", v.rows_b),
        ("キー一致", v.matched),
        ("完全一致", v.same),
        ("値差異", v.changed),
        ("未投入(Aのみ)", v.only_a),
        ("想定外(Bのみ)", v.only_b),
        ("照合不能(A: キー重複・空)", v.unmatchable_a),
        ("照合不能(B: キー重複・空)", v.unmatchable_b),
    ]:
        c = WriteOnlyCell(ws, label)
        c.font = bold
        ws.append([c, value])
    ws.append([])
    for p in v.problems:
        ws.append([p])
    ws.append([])
    ws.append(["ファイルA", diff.name_a])
    ws.append(["ファイルB", diff.name_b])

    table = build_verification_table(diff, only_b_is_error=only_b_is_error)
    ws2 = wb.create_sheet("問題レコード")
    for i, w in enumerate(_col_widths(table.columns, table.rows), start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    head = []
    for vcol in table.columns:
        c = WriteOnlyCell(ws2, vcol)
        c.font = bold
        c.fill = fill_header
        head.append(c)
    ws2.append(head)
    for row in table.rows:
        cells = [WriteOnlyCell(ws2, x) for x in row]
        fill = fills.get(row[0])
        if fill:
            cells[0].fill = fill
        ws2.append(cells)

    # 監査証跡: 手動紐づけ・既知差分の登録内容(誰が見ても判断の記録を追えるように)
    from .workspace import load_known_diffs, load_manual_links

    def _audit_sheet(title, columns, rows):
        ws_a = wb.create_sheet(title)
        head_a = []
        for col in columns:
            c = WriteOnlyCell(ws_a, col)
            c.font = bold
            c.fill = fill_header
            head_a.append(c)
        ws_a.append(head_a)
        for r in rows:
            ws_a.append(list(r))

    links = load_manual_links()
    _audit_sheet(
        "手動紐づけ",
        ["基準(A)キー", "比較(B)キー", "一致率", "メモ", "登録日時"],
        [["/".join(e["key_a"]), "/".join(e["key_b"]),
          "" if e.get("score") is None else e["score"],
          e.get("note", ""), e.get("added_at", "")] for e in links])
    knowns = load_known_diffs()
    _audit_sheet(
        "既知差分",
        ["種類", "キー", "内容", "メモ", "登録日時"],
        [["セル差分" if e.get("type") == "cell" else "欠落",
          "/".join(e.get("key", [])),
          (f"{e.get('col_a', '')}: {e.get('value_a', '')} → {e.get('value_b', '')}"
           if e.get("type") == "cell"
           else ("比較先に無い(未登録を容認)" if e.get("status") == "only_a"
                 else "基準に無い(存在を容認)")),
          e.get("note", ""), e.get("added_at", "")] for e in knowns])

    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()
