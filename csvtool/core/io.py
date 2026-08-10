"""ファイル入出力: エンコーディング/区切り文字の自動判定、CSV/xlsxの読み書き。

方針:
- 値は全て文字列として扱う(先頭ゼロや空セルを壊さない)。
- エンコーディング判定は BOM → UTF-8厳密 → CP932厳密 → charset-normalizer の優先順。
- Windows拡張文字(①㈱～等)のため shift_jis ではなく cp932 を使う。
"""
from __future__ import annotations

import csv
import io as _io
from datetime import date, datetime, time

from .model import CsvToolError, Table

SUPPORTED_ENCODINGS = ("utf-8-sig", "utf-8", "cp932", "utf-16", "euc_jp")
SUPPORTED_DELIMITERS = (",", "\t", ";", "|")

EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xltx")


class EncodingWriteError(CsvToolError):
    """指定エンコーディングで表現できない文字がある場合のエラー。

    details["locations"] に {"row": 1始まり(0=ヘッダー), "column": 列名, "char": 文字}
    のリスト(最大20件)を持つ。
    """

    code = "encoding_write_error"


def is_excel_filename(filename: str) -> bool:
    return filename.lower().endswith(EXCEL_EXTENSIONS)


# ---------------------------------------------------------------- encoding

def detect_encoding(raw: bytes) -> tuple[str, float]:
    """バイト列からエンコーディングを推定して (encoding, confidence) を返す。"""
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig", 1.0
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return "utf-16", 1.0
    sample = raw[:1_000_000]
    try:
        sample.decode("utf-8")
        return "utf-8", 1.0 if sample else 0.5
    except UnicodeDecodeError:
        pass
    try:
        sample.decode("cp932")
        return "cp932", 0.9
    except UnicodeDecodeError:
        pass
    try:
        from charset_normalizer import from_bytes
        best = from_bytes(sample).best()
        if best and best.encoding:
            return best.encoding, 0.5
    except Exception:
        pass
    return "utf-8", 0.0


def decode_bytes(raw: bytes, encoding: str) -> str:
    try:
        return raw.decode(encoding)
    except (UnicodeDecodeError, LookupError) as e:
        raise CsvToolError(
            f"エンコーディング {encoding} でデコードできません: {e}", encoding=encoding
        )


# ---------------------------------------------------------------- delimiter

def detect_delimiter(text: str) -> str:
    """先頭数十行から区切り文字を推定する。判定できなければカンマ。"""
    lines = [ln for ln in text.splitlines() if ln.strip()][:30]
    if not lines:
        return ","
    sample = "\n".join(lines)
    best_delim, best_score = ",", -1.0
    for d in SUPPORTED_DELIMITERS:
        try:
            counts = [len(row) for row in csv.reader(_io.StringIO(sample), delimiter=d)]
        except csv.Error:
            continue
        counts = [c for c in counts if c > 0]
        if not counts:
            continue
        width = max(set(counts), key=counts.count)  # 最頻の列数
        if width < 2:
            continue
        consistency = counts.count(width) / len(counts)
        score = consistency * width
        if score > best_score:
            best_delim, best_score = d, score
    return best_delim


# ---------------------------------------------------------------- headers

def _dedupe_headers(headers: list[str]) -> list[str]:
    """空ヘッダーを「列N」で補い、重複には _2, _3 を付けて一意化する。"""
    result: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(headers):
        name = h.strip() if h else ""
        if not name:
            name = f"列{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
            while name in seen:
                name += "_x"
        seen[name] = 1
        result.append(name)
    return result


def _rows_to_table(all_rows: list[list[str]], header_row: int, name: str) -> Table:
    if header_row < 0 or header_row >= max(len(all_rows), 1):
        raise CsvToolError(
            f"ヘッダー行の指定が範囲外です: {header_row + 1}行目 (全{len(all_rows)}行)",
            header_row=header_row,
        )
    if not all_rows:
        return Table(columns=[], rows=[], name=name)
    headers = _dedupe_headers([str(c) for c in all_rows[header_row]])
    data = [[str(c) for c in row] for row in all_rows[header_row + 1:]]
    t = Table(columns=headers, rows=data, name=name)
    t.make_rectangular()
    return t


# ---------------------------------------------------------------- CSV read

def load_csv(raw: bytes, *, name: str = "", encoding: str | None = None,
             delimiter: str | None = None, header_row: int = 0) -> Table:
    enc = encoding or detect_encoding(raw)[0]
    text = decode_bytes(raw, enc)
    delim = delimiter or detect_delimiter(text)
    if delim not in SUPPORTED_DELIMITERS:
        raise CsvToolError(f"未対応の区切り文字です: {delim!r}")
    reader = csv.reader(_io.StringIO(text), delimiter=delim)
    all_rows = [row for row in reader]
    # 末尾の完全空行は落とす
    while all_rows and all(c == "" for c in all_rows[-1]):
        all_rows.pop()
    table = _rows_to_table(all_rows, header_row, name)
    table.source_encoding = enc
    return table


# ---------------------------------------------------------------- Excel read

def _excel_cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == value.minute == value.second == 0 and value.microsecond == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


def list_sheets(raw: bytes) -> list[str]:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise CsvToolError(f"Excelファイルを開けません: {e}")
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def load_excel(raw: bytes, *, name: str = "", sheet: str | None = None,
               header_row: int = 0) -> Table:
    from openpyxl import load_workbook
    try:
        wb = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise CsvToolError(f"Excelファイルを開けません: {e}")
    try:
        if sheet is None:
            ws = wb[wb.sheetnames[0]]
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            raise CsvToolError(f"シートが見つかりません: {sheet}", sheet=sheet)
        all_rows = [[_excel_cell_to_str(c) for c in row]
                    for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    while all_rows and all(c == "" for c in all_rows[-1]):
        all_rows.pop()
    table = _rows_to_table(all_rows, header_row, name)
    table.sheet = ws.title
    return table


def load_table(raw: bytes, filename: str, *, encoding: str | None = None,
               delimiter: str | None = None, sheet: str | None = None,
               header_row: int = 0) -> Table:
    """拡張子でCSV/Excelを判別して読み込む。"""
    if is_excel_filename(filename):
        return load_excel(raw, name=filename, sheet=sheet, header_row=header_row)
    return load_csv(raw, name=filename, encoding=encoding,
                    delimiter=delimiter, header_row=header_row)


# ---------------------------------------------------------------- write

def _find_unencodable(table: Table, encoding: str, limit: int = 20) -> list[dict]:
    locations: list[dict] = []
    for ci, col in enumerate(table.columns):
        try:
            col.encode(encoding)
        except UnicodeEncodeError as e:
            locations.append({"row": 0, "column": col, "char": col[e.start:e.end]})
            if len(locations) >= limit:
                return locations
    for ri, row in enumerate(table.rows, start=1):
        for ci, cell in enumerate(row):
            try:
                cell.encode(encoding)
            except UnicodeEncodeError as e:
                locations.append({
                    "row": ri,
                    "column": table.columns[ci] if ci < len(table.columns) else str(ci),
                    "char": cell[e.start:e.end],
                })
                if len(locations) >= limit:
                    return locations
    return locations


def write_csv(table: Table, *, encoding: str = "utf-8-sig", delimiter: str = ",",
              errors: str = "strict") -> bytes:
    """TableをCSVバイト列にする(CRLF)。encodeできない文字は位置付きでエラー報告。"""
    buf = _io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter, lineterminator="\r\n",
                        quoting=csv.QUOTE_MINIMAL)
    writer.writerow(table.columns)
    writer.writerows(table.rows)
    text = buf.getvalue()
    try:
        return text.encode(encoding, errors=errors)
    except UnicodeEncodeError:
        locations = _find_unencodable(table, encoding)
        raise EncodingWriteError(
            f"{encoding} で保存できない文字があります(該当 {len(locations)} 箇所以上)。"
            "エンコーディングを UTF-8 にするか、置換保存を選んでください。",
            locations=locations, encoding=encoding,
        )


def write_xlsx(table: Table, *, sheet_name: str = "Sheet1") -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    ws.append(table.columns)
    for row in table.rows:
        ws.append(row)
    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()
