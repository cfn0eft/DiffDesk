"""色付きExcel差分レポートの生成(openpyxl)。"""
from __future__ import annotations

import io as _io

from .model import DiffResult
from .upsert import build_report_table

_FILL_ONLY_A = "C6EFCE"   # 緑: Aのみ(insert候補)
_FILL_ONLY_B = "FFC7CE"   # 赤: Bのみ
_FILL_CHANGED = "FFEB9C"  # 黄: 変更セル
_FILL_HEADER = "D9E1F2"


def build_xlsx_report(diff: DiffResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # --- サマリーシート
    ws = wb.active
    ws.title = "サマリー"
    s = diff.summary
    ws.append(["項目", "件数"])
    for label, key in [("Aのみ(insert候補)", "only_a"), ("Bのみ", "only_b"),
                       ("変更", "changed"), ("一致", "same"),
                       ("キー重複(A)", "duplicates_a"), ("キー重複(B)", "duplicates_b"),
                       ("キー空(A)", "empty_key_a"), ("キー空(B)", "empty_key_b")]:
        ws.append([label, s[key]])
    ws.append([])
    ws.append(["ファイルA", diff.name_a])
    ws.append(["ファイルB", diff.name_b])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # --- 詳細シート
    report = build_report_table(diff)
    ws2 = wb.create_sheet("差分詳細")
    header_fill = PatternFill("solid", fgColor=_FILL_HEADER)
    fills = {
        "Aのみ": PatternFill("solid", fgColor=_FILL_ONLY_A),
        "Bのみ": PatternFill("solid", fgColor=_FILL_ONLY_B),
    }
    changed_fill = PatternFill("solid", fgColor=_FILL_CHANGED)

    ws2.append(report.columns)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # 変更セルの位置を特定するため、値ペア列の開始位置を計算
    n_keys = len(diff.mapping.key_pairs)
    value_pairs = diff.mapping.value_pairs

    for rd, report_row in zip(diff.rows, report.rows):
        ws2.append(report_row)
        r = ws2.max_row
        status_cell = ws2.cell(row=r, column=1)
        if report_row[0] in fills:
            for c in range(1, len(report.columns) + 1):
                ws2.cell(row=r, column=c).fill = fills[report_row[0]]
        elif report_row[0] == "変更":
            status_cell.fill = changed_fill
            changed_cols = {cd.col_a for cd in rd.cell_diffs}
            for vi, p in enumerate(value_pairs):
                if p.col_a in changed_cols:
                    base = 1 + n_keys + vi * 2 + 1  # 状態+キー列の後、A/Bペア
                    ws2.cell(row=r, column=base).fill = changed_fill
                    ws2.cell(row=r, column=base + 1).fill = changed_fill

    ws2.freeze_panes = "A2"
    for ws_ in (ws, ws2):
        for col_cells in ws_.columns:
            width = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws_.column_dimensions[col_cells[0].column_letter].width = min(width * 1.6 + 2, 40)

    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()
