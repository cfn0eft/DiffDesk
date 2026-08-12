import pytest
from fastapi.testclient import TestClient

import diffdesk.core.profile as profile_mod
from diffdesk.web.app import create_app
from diffdesk.web.sessions import SessionStore
from tests.conftest import load_fixture


@pytest.fixture
def client(monkeypatch, tmp_path):
    monkeypatch.setattr(profile_mod, "DEFAULT_PROFILE_DIR", tmp_path / "profiles")
    # ストアをテストごとに初期化
    import diffdesk.web.routes as routes_mod
    import diffdesk.web.sessions as sessions_mod
    fresh = SessionStore()
    monkeypatch.setattr(sessions_mod, "store", fresh)
    monkeypatch.setattr(routes_mod, "store", fresh)
    return TestClient(create_app())


def upload(client, name: str, content: bytes) -> dict:
    r = client.post("/api/files", files={"file": (name, content)})
    assert r.status_code == 200, r.text
    return r.json()


MAPPING = {"pairs": [
    {"col_a": "社員番号", "col_b": "EmployeeNumber__c", "is_key": True},
    {"col_a": "氏名", "col_b": "Name"},
    {"col_a": "メール", "col_b": "Email"},
    {"col_a": "部署", "col_b": "Department__c", "sf_field": "Busho__c"},
]}


class TestFiles:
    def test_upload_detects_cp932(self, client):
        info = upload(client, "master.csv", load_fixture("master_cp932.csv"))
        assert info["detected_encoding"] == "cp932"
        assert info["preview"]["columns"][0] == "氏名"
        assert info["preview"]["rows"][0][3] == "0001"  # 先頭ゼロ保持

    def test_upload_excel_lists_sheets(self, client):
        info = upload(client, "master.xlsx", load_fixture("master.xlsx"))
        assert info["sheets"] == ["社員マスタ", "タイトル付き"]
        assert info["preview"]["columns"][0] == "氏名"

    def test_reparse_with_sheet_and_header(self, client):
        info = upload(client, "master.xlsx", load_fixture("master.xlsx"))
        r = client.post(f"/api/files/{info['file_id']}/parse",
                        json={"sheet": "タイトル付き", "header_row": 2})
        assert r.status_code == 200
        assert r.json()["preview"]["columns"][0] == "氏名"

    def test_reparse_bad_encoding_400(self, client):
        info = upload(client, "master.csv", load_fixture("master_cp932.csv"))
        r = client.post(f"/api/files/{info['file_id']}/parse",
                        json={"encoding": "utf-8"})
        assert r.status_code == 400
        assert "エラー" not in r.json()["error"]["code"]  # 構造化エラー

    def test_get_paged(self, client):
        info = upload(client, "m.csv", load_fixture("master_utf8.csv"))
        r = client.get(f"/api/files/{info['file_id']}?offset=1&limit=2")
        body = r.json()
        assert body["total_rows"] == 5 and len(body["rows"]) == 2

    def test_put_table_and_export(self, client):
        info = upload(client, "m.csv", load_fixture("master_utf8.csv"))
        fid = info["file_id"]
        r = client.put(f"/api/files/{fid}/table",
                       json={"columns": ["a"], "rows": [["あ"]]})
        assert r.status_code == 200
        r = client.post(f"/api/export/table/{fid}",
                        json={"encoding": "cp932", "filename": "編集済"})
        assert r.status_code == 200
        assert r.content.decode("cp932") == "a\r\nあ\r\n"
        assert "%E7%B7%A8%E9%9B%86%E6%B8%88" in r.headers["content-disposition"]

    def test_export_unknown_encoding_400(self, client):
        info = upload(client, "m.csv", load_fixture("master_utf8.csv"))
        r = client.post(f"/api/export/table/{info['file_id']}",
                        json={"encoding": "nope-encoding"})
        assert r.status_code == 400
        assert "エンコーディング" in r.json()["error"]["message"]

    def test_upload_xls_clear_error(self, client):
        ole = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 16
        r = client.post("/api/files", files={"file": ("old.xls", ole)})
        assert r.status_code == 400
        assert ".xlsx" in r.json()["error"]["message"]

    def test_export_unencodable_400_with_locations(self, client):
        info = upload(client, "m.csv", "col\n𩸽\n".encode("utf-8"))
        r = client.post(f"/api/export/table/{info['file_id']}",
                        json={"encoding": "cp932"})
        assert r.status_code == 400
        assert r.json()["error"]["locations"][0]["row"] == 1

    def test_clean_validate_dedupe_replace(self, client):
        raw = "名前,日付\nＡＢＣ,2020/1/5\nＡＢＣ,2020/1/5\n".encode("utf-8")
        fid = upload(client, "x.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/dedupe")
        assert r.json()["removed"] == 1
        r = client.post(f"/api/files/{fid}/clean",
                        json={"columns": ["名前", "日付"], "ops": ["zen2han", "date_iso"]})
        assert r.json()["changed_cells"] == 2
        r = client.post(f"/api/files/{fid}/replace",
                        json={"query": "ABC", "replacement": "XYZ"})
        assert r.json()["replaced"] == 1
        r = client.post(f"/api/files/{fid}/validate",
                        json={"required_columns": ["名前"]})
        assert r.json()["count"] == 0

    def test_concat_and_enrich(self, client):
        f1 = upload(client, "a.csv", "id,v\n1,x\n".encode("utf-8"))["file_id"]
        f2 = upload(client, "b.csv", "id,w\n1,y\n".encode("utf-8"))["file_id"]
        r = client.post("/api/files/concat", json={"file_ids": [f1, f2]})
        assert r.status_code == 200
        assert r.json()["preview"]["total_rows"] == 2
        r = client.post(f"/api/files/{f1}/enrich",
                        json={"other_file_id": f2,
                              "key_pairs": [["id", "id"]], "add_columns": ["w"]})
        body = r.json()
        assert body["preview"]["columns"] == ["id", "v", "w"]
        assert body["unmatched"] == 0


class TestDiffFlow:
    def run_diff(self, client) -> tuple[str, dict]:
        fa = upload(client, "master.csv", load_fixture("master_cp932.csv"))["file_id"]
        fb = upload(client, "sf.csv", load_fixture("salesforce_export.csv"))["file_id"]
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb, "mapping": MAPPING,
        })
        assert r.status_code == 200, r.text
        body = r.json()
        return body["diff_id"], body

    def test_automap(self, client):
        fa = upload(client, "a.csv", "Name,Email\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv", "ｎａｍｅ,Email,Extra\n".encode("utf-8"))["file_id"]
        r = client.post("/api/automap", json={"file_a": fa, "file_b": fb})
        body = r.json()
        assert body["by_name"] == 2
        assert any(p["col_a"] == "Email" and p["col_b"] == "Email" for p in body["pairs"])
        assert any(p["col_a"] == "Name" and p["col_b"] == "ｎａｍｅ" for p in body["pairs"])

    def test_automap_value_based(self, client):
        fa = upload(client, "master.csv", load_fixture("master_utf8.csv"))["file_id"]
        fb = upload(client, "sf.csv", load_fixture("salesforce_export.csv"))["file_id"]
        r = client.post("/api/automap", json={"file_a": fa, "file_b": fb})
        body = r.json()
        by = {p["col_a"]: p["col_b"] for p in body["pairs"]}
        # 名前が一致しない日本語↔SF API名でも辞書・値ベースで対応付けられる
        assert by["社員番号"] == "EmployeeNumber__c"
        assert by["氏名"] == "Name"
        assert by["メール"] == "Email"

    def test_verify_endpoint_and_export(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        v = client.get(f"/api/diff/{diff_id}/verify").json()
        assert v["passed"] is False and v["only_a"] == 2
        v2 = client.get(f"/api/diff/{diff_id}/verify?only_b_is_error=false").json()
        assert v2["only_b"] == 1 and v2["passed"] is False  # 未投入があるため依然NG
        r = client.post(f"/api/export/verify/{diff_id}", json={"format": "csv"})
        text = r.content.decode("utf-8-sig")
        assert len(text.strip().splitlines()) == 6  # ヘッダー+問題5行
        r = client.post(f"/api/export/verify/{diff_id}", json={"format": "xlsx"})
        assert r.content[:2] == b"PK"

    def test_diff_summary_and_rows(self, client):
        diff_id, body = self.run_diff(client)
        assert body["summary"]["changed"] == 2
        r = client.get(f"/api/diff/{diff_id}/rows?status=changed")
        rows = r.json()["rows"]
        assert len(rows) == 2
        assert rows[0]["cell_diffs"]

    def test_export_upsert(self, client):
        diff_id, _ = self.run_diff(client)
        r = client.post(f"/api/export/upsert/{diff_id}",
                        json={"external_id": "社員番号"})
        assert r.status_code == 200
        text = r.content.decode("utf-8-sig")
        assert text.splitlines()[0] == "EmployeeNumber__c,Name,Email,Busho__c"
        assert len(text.strip().splitlines()) == 5  # ヘッダー+4行

    def test_export_upsert_nothing_selected_400(self, client):
        diff_id, _ = self.run_diff(client)
        r = client.post(f"/api/export/upsert/{diff_id}",
                        json={"external_id": "社員番号",
                              "include_insert": False, "include_update": False})
        assert r.status_code == 400  # 全件出力ではなく明示エラー

    def test_export_delete_requires_id_mapping(self, client):
        diff_id, _ = self.run_diff(client)
        r = client.post(f"/api/export/delete/{diff_id}", json={})
        assert r.status_code == 400  # Id列がマッピングにない

    def test_export_sdl_and_reports(self, client):
        diff_id, _ = self.run_diff(client)
        assert "Busho__c=Busho__c" in client.get(f"/api/export/sdl/{diff_id}").text
        r = client.post(f"/api/export/report/{diff_id}", json={"format": "csv"})
        assert "状態" in r.content.decode("utf-8-sig")
        r = client.post(f"/api/export/report/{diff_id}", json={"format": "xlsx"})
        assert r.content[:2] == b"PK"

    def test_merge(self, client):
        diff_id, _ = self.run_diff(client)
        r = client.post(f"/api/diff/{diff_id}/merge", json={
            "choices": [{"key": ["0002"], "col_a": "メール", "use": "b"}],
        })
        body = r.json()
        row = next(row for row in body["preview"]["rows"] if row[0] == "0002")
        assert row[2] == "hanako-new@example.com"

    def test_diff_error_missing_key(self, client):
        fa = upload(client, "a.csv", "x\n1\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv", "y\n1\n".encode("utf-8"))["file_id"]
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb,
            "mapping": {"pairs": [{"col_a": "x", "col_b": "y"}]},
        })
        assert r.status_code == 400
        assert "キー" in r.json()["error"]["message"]


class TestV040Features:
    def test_clusters_and_apply(self, client):
        raw = "会社\n株式会社テスト\nテスト(株)\n別会社\n".encode("utf-8")
        fid = upload(client, "c.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/clusters", json={"column": "会社"}).json()
        assert r["count"] == 1
        mapping = {"テスト(株)": "株式会社テスト"}
        r = client.post(f"/api/files/{fid}/apply-map",
                        json={"column": "会社", "mapping": mapping}).json()
        assert r["changed"] == 1

    def test_error_analysis_and_retry(self, client):
        raw = ("氏名,ERROR\n山田,REQUIRED_FIELD_MISSING:必須項目\n"
               "鈴木,UNABLE_TO_LOCK_ROW:ロック\n").encode("utf-8")
        fid = upload(client, "error.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/analyze-errors").json()
        assert r["total_rows"] == 2 and len(r["categories"]) == 2
        res = client.post(f"/api/export/retry/{fid}", json={})
        text = res.content.decode("utf-8-sig")
        assert text.splitlines()[0] == "氏名"  # ERROR列除去
        assert len(text.strip().splitlines()) == 3

    def test_anonymize_creates_new_file(self, client):
        fid = upload(client, "m.csv", load_fixture("master_utf8.csv"))["file_id"]
        r = client.post(f"/api/files/{fid}/anonymize",
                        json={"spec": {"氏名": "name", "メール": "email"}}).json()
        assert r["changed"] == 10
        assert r["file_id"] != fid
        assert "@example.com" in r["preview"]["rows"][0][1]

    def test_split_column(self, client):
        raw = "氏名\n山田 太郎\n鈴木 花子\n".encode("utf-8")
        fid = upload(client, "s.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/split-column",
                        json={"column": "氏名", "delimiter": " "}).json()
        assert r["parts"] == 2
        assert r["preview"]["columns"] == ["氏名_1", "氏名_2"]

    def test_column_values_and_allowed_validation(self, client):
        fid = upload(client, "m.csv", load_fixture("master_utf8.csv"))["file_id"]
        r = client.post(f"/api/files/{fid}/column-values", json={"column": "部署"}).json()
        assert set(r["values"]) == {"営業部", "開発部", "総務部"}
        v = client.post(f"/api/files/{fid}/validate",
                        json={"allowed_values": {"部署": ["営業部", "開発部"]}}).json()
        assert v["count"] == 1  # 総務部が許可外

    def test_clean_ops_includes_fill_down(self, client):
        ops = {o["id"] for o in client.get("/api/clean-ops").json()["ops"]}
        assert {"fill_down", "digits_only", "alnum_han"} <= ops


class TestV050Features:
    def test_restore_export(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.post(f"/api/export/restore/{diff_id}", json={})
        text = r.content.decode("utf-8-sig")
        assert "hanako-new@example.com" in text  # 投入前のB値
        assert len(text.strip().splitlines()) == 3  # ヘッダー+changed2行

    def test_undo_delete(self, client):
        raw = ("ID,STATUS\na01x,Item Created\na01y,Item Updated\n").encode("utf-8")
        fid = upload(client, "success.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/undo-delete", json={})
        assert r.status_code == 200
        text = r.content.decode("utf-8-sig")
        assert text.strip().splitlines() == ["Id", "a01x"]
        assert r.headers["X-Skipped-Updates"] == "1"

    def test_html_report(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.post(f"/api/export/html/{diff_id}", json={})
        html_text = r.content.decode("utf-8")
        assert "<!DOCTYPE html>" in html_text and "要確認" in html_text

    def test_profile_and_baseline(self, client):
        fid = upload(client, "m.csv", load_fixture("master_utf8.csv"))["file_id"]
        p = client.post(f"/api/files/{fid}/profile", json={}).json()
        assert p["rows"] == 5
        client.post(f"/api/files/{fid}/save-baseline", json={"name": "月次"})
        assert client.get("/api/baselines").json()["baselines"] == ["月次"]
        w = client.post(f"/api/files/{fid}/compare-baseline",
                        json={"name": "月次"}).json()["warnings"]
        assert w[0]["level"] == "info"  # 同一ファイルなので変化なし

    def test_fuzzy_match_and_link(self, client):
        fa = upload(client, "a.csv", "氏名\n山田太郎\n鈴木花子\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv", "name\n山田 太郎\n別人\n".encode("utf-8"))["file_id"]
        r = client.post("/api/fuzzy-match", json={
            "file_a": fa, "file_b": fb, "pairs": [["氏名", "name"]], "threshold": 0.7,
        }).json()
        assert r["count"] == 1
        assert r["candidates"][0]["values_a"] == ["山田太郎"]
        r2 = client.post("/api/fuzzy-link", json={
            "file_a": fa, "file_b": fb, "matches": r["candidates"],
        }).json()
        assert r2["preview"]["columns"] == ["氏名", "name", "突合スコア"]

    def test_crosstab(self, client):
        raw = "群,性別\nG1,雄\nG1,雌\nG2,雄\n".encode("utf-8")
        fid = upload(client, "g.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/crosstab",
                        json={"row_col": "群", "col_col": "性別"}).json()
        assert r["table"]["columns"] == ["群\\性別", "雄", "雌", "(合計)"]

    def test_split_address(self, client):
        raw = "住所\n東京都港区芝1-2-3\n".encode("utf-8")
        fid = upload(client, "addr.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/split-address", json={"column": "住所"}).json()
        assert r["parsed"] == 1
        assert "住所_都道府県" in r["preview"]["columns"]

    def test_validate_ranges(self, client):
        raw = "ALT\n25\n70\n".encode("utf-8")
        fid = upload(client, "lab.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/validate",
                        json={"ranges": {"ALT": [10, 50]}}).json()
        assert r["count"] == 1 and "範囲外" in r["issues"][0]["message"]


class TestV060Features:
    def test_calc_column_and_recipe_roundtrip(self, client, monkeypatch, tmp_path):
        raw = "姓,名,部署\n山田, ＡＢ ,営業部\n鈴木,子,開発部\n".encode("utf-8")
        fid = upload(client, "r.csv", raw)["file_id"]
        # クレンジング → 計算列(結合) → 履歴確認 → レシピ保存 → 別ファイルに適用
        client.post(f"/api/files/{fid}/clean",
                    json={"columns": ["名"], "ops": ["trim", "zen2han"]})
        r = client.post(f"/api/files/{fid}/calc-column", json={
            "mode": "concat", "new_name": "氏名", "columns": ["姓", "名"],
            "separator": " "})
        assert "氏名" in r.json()["preview"]["columns"]
        rec = client.get(f"/api/files/{fid}/recipe").json()
        assert len(rec["ops"]) == 2
        assert any("クレンジング" in l for l in rec["labels"])
        client.post("/api/recipes", json={"name": "整形テスト", "file_id": fid})
        assert "整形テスト" in client.get("/api/recipes").json()["recipes"]

        fid2 = upload(client, "r2.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid2}/apply-recipe", json={"name": "整形テスト"})
        body = r.json()
        assert len(body["logs"]) == 2
        assert "氏名" in body["preview"]["columns"]
        assert body["preview"]["rows"][0][3] == "山田 AB"

    def test_conditional_calc(self, client):
        raw = "点数\n80\n40\n".encode("utf-8")
        fid = upload(client, "s.csv", raw)["file_id"]
        r = client.post(f"/api/files/{fid}/calc-column", json={
            "mode": "conditional", "new_name": "判定", "column": "点数",
            "op": "gte", "value": "60", "then_value": "合格", "else_value": "不合格"})
        rows = r.json()["preview"]["rows"]
        assert rows[0][1] == "合格" and rows[1][1] == "不合格"


class TestV070Features:
    def test_known_diff_flow(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        # 変更セル(0002のメール)を既知に登録
        r = client.post("/api/known-diffs", json={"entry": {
            "type": "cell", "key": ["0002"], "col_a": "メール",
            "value_a": "hanako@example.com", "value_b": "hanako-new@example.com"}})
        assert r.status_code == 200
        # rowsに反映(0002はsame+known_diffsになる)
        rows = client.get(f"/api/diff/{diff_id}/rows").json()["rows"]
        r2 = next(x for x in rows if x["key"] == ["0002"])
        assert r2["status"] == "same" and len(r2["known_diffs"]) == 1
        # 検証にも反映
        v = client.get(f"/api/diff/{diff_id}/verify").json()
        assert v["changed"] == 1 and v["known_cells"] == 1
        # 列サマリー(メールの差異が消えて部署のみ)
        cs = client.get(f"/api/diff/{diff_id}/columns-summary").json()["columns"]
        assert cs == [{"column": "部署", "count": 1}]
        # 管理APIで削除
        client.delete("/api/known-diffs")
        v = client.get(f"/api/diff/{diff_id}/verify").json()
        assert v["changed"] == 2

    def test_history_recorded(self, client):
        TestDiffFlow().run_diff(client)
        h = client.get("/api/history").json()["history"]
        assert len(h) == 1 and h[0]["name_a"] == "master.csv"
        client.delete("/api/history")
        assert client.get("/api/history").json()["history"] == []

    def test_user_dict_automap(self, client):
        fa = upload(client, "a.csv", "検体番号\nS-1\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv", "SampleNo__c,Other__c\nS-1,S-1\n".encode("utf-8"))["file_id"]
        client.post("/api/user-dict",
                    json={"pairs": [{"col_a": "検体番号", "col_b": "SampleNo__c"}]})
        r = client.post("/api/automap", json={"file_a": fa, "file_b": fb}).json()
        assert r["by_dict"] == 1
        by = {p["col_a"]: p["col_b"] for p in r["pairs"]}
        assert by["検体番号"] == "SampleNo__c"


class TestProfiles:
    def test_save_load_list_delete(self, client):
        profile = {
            "name": "月次",
            "mapping": MAPPING,
            "options": {"ignore_case": True},
            "external_id": "社員番号",
        }
        r = client.post("/api/profiles", json={"profile": profile})
        assert r.status_code == 200
        assert client.get("/api/profiles").json()["profiles"] == ["月次"]
        loaded = client.get("/api/profiles/月次").json()["profile"]
        assert loaded["mapping"]["pairs"][0]["is_key"]
        assert loaded["options"]["ignore_case"]
        client.delete("/api/profiles/月次")
        assert client.get("/api/profiles").json()["profiles"] == []


def test_index_served(client):
    r = client.get("/")
    assert r.status_code == 200
    assert "html" in r.text


class TestV080Features:
    def test_manual_pair_flow(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        # 紐づけ相手の候補一覧
        ua = client.get(f"/api/diff/{diff_id}/unmatched", params={"side": "a"}).json()["rows"]
        ub = client.get(f"/api/diff/{diff_id}/unmatched", params={"side": "b"}).json()["rows"]
        assert [x["key"] for x in ua] == [["0004"], ["0005"]]
        assert [x["key"] for x in ub] == [["0006"]]
        # 手動で紐づけ
        r = client.post(f"/api/diff/{diff_id}/manual-pairs",
                        json={"key_a": ["0004"], "key_b": ["0006"]})
        assert r.status_code == 200 and r.json()["count"] == 1
        rows = client.get(f"/api/diff/{diff_id}/rows").json()["rows"]
        m = next(x for x in rows if x["key"] == ["0004"])
        assert m["manual"] and m["key_b"] == ["0006"] and m["status"] == "changed"
        assert not any(x["status"] == "only_b" for x in rows)
        v = client.get(f"/api/diff/{diff_id}/verify").json()
        assert v["only_a"] == 1 and v["only_b"] == 0
        # 候補一覧からも消える
        ua = client.get(f"/api/diff/{diff_id}/unmatched", params={"side": "a"}).json()["rows"]
        assert [x["key"] for x in ua] == [["0005"]]
        # 解除で元に戻る
        client.delete(f"/api/diff/{diff_id}/manual-pairs/0")
        v = client.get(f"/api/diff/{diff_id}/verify").json()
        assert v["only_a"] == 2 and v["only_b"] == 1

    def test_manual_pair_validation(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        # キー一致済みの行は紐づけ不可
        r = client.post(f"/api/diff/{diff_id}/manual-pairs",
                        json={"key_a": ["0001"], "key_b": ["0006"]})
        assert r.status_code == 400
        # 存在しないキー
        r = client.post(f"/api/diff/{diff_id}/manual-pairs",
                        json={"key_a": ["0004"], "key_b": ["9999"]})
        assert r.status_code == 400


class TestV081Serving:
    def test_index_no_store_and_version_injected(self, client):
        r = client.get("/")
        assert r.headers["cache-control"] == "no-store"
        from diffdesk import __version__
        assert f"v{__version__}" in r.text
        assert "{{V}}" not in r.text  # プレースホルダが残っていない

    def test_static_requires_revalidation(self, client):
        r = client.get("/static/app.js")
        assert r.status_code == 200
        assert r.headers["cache-control"] == "no-cache"


class TestV090Features:
    def test_pair_preview(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.post(f"/api/diff/{diff_id}/pair-preview",
                        json={"key_a": ["0004"], "key_b": ["0006"]})
        assert r.status_code == 200
        body = r.json()
        assert 0 <= body["score"] <= 1
        assert len(body["columns"]) == 4
        assert {"col_a", "value_a", "value_b", "sim"} <= set(body["columns"][0])

    def test_link_suggest_endpoint(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.post(f"/api/diff/{diff_id}/link-suggest",
                        json={"threshold": 0.1, "limit": 10})
        assert r.status_code == 200
        cands = r.json()["candidates"]
        assert all(c["score"] >= 0.1 for c in cands)

    def test_link_prompt_and_import(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.get(f"/api/diff/{diff_id}/link-prompt")
        assert r.status_code == 200
        assert r.json()["rows_a"] == 2 and r.json()["rows_b"] == 1
        assert "JSON" in r.json()["prompt"]
        # 回答取り込み(有効1件+無効1件)
        text = '回答: ```json\n[{"key_a": ["0004"], "key_b": ["0006"], "confidence": 0.9, "reason": "類似"}, {"key_a": ["xxxx"], "key_b": ["0006"]}]\n```'
        r = client.post(f"/api/diff/{diff_id}/link-import", json={"text": text})
        assert r.status_code == 200
        body = r.json()
        assert len(body["pairs"]) == 1 and len(body["rejected"]) == 1
        assert body["pairs"][0]["score"] == 0.9

    def test_manual_link_persists_across_rerun(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.post(f"/api/diff/{diff_id}/manual-pairs",
                        json={"key_a": ["0004"], "key_b": ["0006"],
                              "note": "監査メモ", "score": 0.8})
        assert r.status_code == 200
        # 同じファイルで差分を再実行 → 新しいdiff_idでも自動で再適用される
        diff_id2, _ = TestDiffFlow().run_diff(client)
        assert diff_id2 != diff_id
        rows = client.get(f"/api/diff/{diff_id2}/rows").json()["rows"]
        m = next(x for x in rows if x["key"] == ["0004"])
        assert m["manual"] and m["key_b"] == ["0006"]
        pairs = client.get(f"/api/diff/{diff_id2}/manual-pairs").json()["pairs"]
        assert pairs[0]["note"] == "監査メモ" and pairs[0]["score"] == 0.8

    def test_ranked_unmatched(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        import json as _json
        r = client.get(f"/api/diff/{diff_id}/unmatched",
                       params={"side": "b", "rank_for": _json.dumps(["0004"])})
        rows = r.json()["rows"]
        assert rows and "score" in rows[0]

    def test_verify_xlsx_has_audit_sheets(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        client.post(f"/api/diff/{diff_id}/manual-pairs",
                    json={"key_a": ["0004"], "key_b": ["0006"], "note": "test"})
        r = client.post(f"/api/export/verify/{diff_id}", json={"format": "xlsx"})
        assert r.status_code == 200
        import io as _io

        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        assert "手動紐づけ" in wb.sheetnames and "既知差分" in wb.sheetnames
        ws = wb["手動紐づけ"]
        rows = list(ws.values)
        assert rows[0][0] == "基準(A)キー"
        assert rows[1][0] == "0004" and rows[1][3] == "test"


class TestV0100Junction:
    def _setup(self, client):
        src = ("会社コード,製品名,数量\n"
               "C001,製品 X,10\nC001,製品 X,20\nC001,製品Y,5\nC002,製品 X,3\n").encode("utf-8")
        fa = upload(client, "src.csv", src)["file_id"]
        fb = upload(client, "a.csv", "CompanyCode__c\nC001\nC002\n".encode("utf-8"))["file_id"]
        fc = upload(client, "b.csv", "ProductKey__c\n製品X\n製品Y\n".encode("utf-8"))["file_id"]
        fj = upload(client, "j.csv",
                    "RelKey__c\nREL-C001-製品X\nREL-C001-製品Y\n".encode("utf-8"))["file_id"]
        config = {
            "a_source_col": "会社コード", "b_source_col": "製品名",
            "a_ext_col": "CompanyCode__c", "b_ext_col": "ProductKey__c",
            "j_key_col": "RelKey__c", "key_template": "REL-{A}-{B}",
            "b_regex_pattern": r"\s+", "b_regex_replacement": "",
        }
        return {"file_source": fa, "file_a": fb, "file_b": fc, "file_j": fj,
                "config": config}

    def test_verify_and_orphans(self, client):
        req = self._setup(client)
        r = client.post("/api/junction-verify", json=req)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["parent_a"]["passed"] and body["parent_b"]["passed"]
        # 中間: 期待3件(C001-X, C001-Y, C002-X)に対し実績2件 → C002-Xが未取込
        assert body["junction"]["expected"] == 3
        assert body["junction"]["missing"] == ["REL-C002-製品X"]
        assert not body["passed"]
        # 両親は存在するので parents_ok
        assert body["orphans"]["causes"]["parents_ok"] == 1
        # 未取込CSV
        r2 = client.post("/api/junction-verify/orphans", json={**req, "encoding": "utf-8"})
        assert r2.status_code == 200
        text = r2.content.decode("utf-8")
        assert "REL-C002-製品X" in text and "中間だけ未取込" in text

    def test_config_validation_error(self, client):
        req = self._setup(client)
        req["config"]["key_template"] = "REL-{A}"  # {B}が無い
        r = client.post("/api/junction-verify", json=req)
        assert r.status_code == 400


class TestV0110MigrationSpec:
    SPEC = {
        "object": "Parent_A__c",
        "externalIdField": "ExtA__c",
        "fields": [
            {"excelColumn": "SrcID", "salesforceField": "ExtA__c", "type": "string"},
            {"excelColumn": "区分", "salesforceField": "Kind__c", "type": "picklist",
             "valueMap": {"甲": "TypeA", "乙": "TypeB"}, "defaultValue": ""},
            {"salesforceField": "IsActive__c", "constantValue": "true", "type": "boolean"},
        ],
    }

    def test_mapping_endpoint(self, client):
        r = client.post("/api/migration-spec/mapping", json={"spec": self.SPEC})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["transform_count"] == 1
        assert body["constants"][0]["field"] == "IsActive__c"
        key = next(p for p in body["pairs"] if p["is_key"])
        assert key["col_b"] == "ExtA__c"

    def test_diff_with_transform(self, client):
        fa = upload(client, "src.csv",
                    "SrcID,区分\nP1,甲\nP2,乙\nP3,不明\n".encode("utf-8"))["file_id"]
        fb = upload(client, "sf.csv",
                    "ExtA__c,Kind__c\nP1,TypeA\nP2,TypeB\nP3,\n".encode("utf-8"))["file_id"]
        pairs = client.post("/api/migration-spec/mapping",
                            json={"spec": self.SPEC}).json()["pairs"]
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb, "mapping": {"pairs": pairs}})
        assert r.status_code == 200, r.text
        # 甲→TypeA、乙→TypeB、マップ外→既定""の変換を再現して全行一致
        assert r.json()["summary"]["same"] == 3

    def test_junction_endpoint(self, client):
        jx = {
            "object": "J__c", "externalIdField": "CompKey__c",
            "skipIfBlank": ["B番号"],
            "fields": [
                {"salesforceField": "CompKey__c", "source": "composite",
                 "prefix": "REL-", "sourceColumns": ["SrcID", "B番号"]},
                {"excelColumn": "SrcID", "salesforceField": "PA__r.ExtA__c",
                 "verifyExternalId": {"object": "Parent_A__c", "field": "ExtA__c"}},
                {"excelColumn": "B番号", "salesforceField": "PB__r.ExtB__c",
                 "verifyExternalId": {"object": "Parent_B__c", "field": "ExtB__c"}},
            ],
        }
        r = client.post("/api/migration-spec/junction", json={"specs": [jx, self.SPEC]})
        assert r.status_code == 200, r.text
        s = r.json()["settings"]
        assert s["key_template"] == "REL-{A}-{B}"
        assert s["j_key_col"] == "CompKey__c" and s["required_col"] == "B番号"

    def test_bad_spec_400(self, client):
        r = client.post("/api/migration-spec/mapping", json={"spec": {"x": 1}})
        assert r.status_code == 400


class TestV0120JunctionSimple:
    def _files(self, client):
        fs = upload(client, "src.csv",
                    "会社,製品\nC1,P1\nC1,P2\nC2,P1\n".encode("utf-8"))["file_id"]
        fj = upload(client, "j.csv",
                    "Key__c\nREL-C1-P1\nREL-C9-P9\n".encode("utf-8"))["file_id"]
        return fs, fj

    def test_two_files_with_relations(self, client):
        fs, fj = self._files(client)
        r = client.post("/api/junction-verify", json={
            "file_source": fs, "file_j": fj,
            "config": {"a_source_col": "会社", "b_source_col": "製品",
                       "j_key_col": "Key__c", "key_template": "REL-{A}-{B}"}})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["parent_a"] is None and body["parent_b"] is None
        assert body["orphans"]["causes"]["unknown"] == 2
        rel = body["relations"]
        by_a = {g["a"]: g for g in rel["groups"]}
        assert by_a["C1"]["ok"] == 1 and by_a["C1"]["ng"] == 1
        assert by_a["C9"]["items"][0]["status"] == "extra"

    def test_infer_template_endpoint(self, client):
        fs, fj = self._files(client)
        r = client.post("/api/junction-verify/infer-template", json={
            "file_source": fs, "file_j": fj,
            "a_source_col": "会社", "b_source_col": "製品",
            "j_key_col": "Key__c"})
        assert r.status_code == 200, r.text
        assert r.json()["template"] == "REL-{A}-{B}"

    def test_file_info_endpoint(self, client):
        fs, _ = self._files(client)
        r = client.get(f"/api/files/{fs}/info")
        assert r.status_code == 200
        body = r.json()
        assert body["filename"] == "src.csv"
        assert body["preview"]["columns"] == ["会社", "製品"]
        assert "parse_params" in body


class TestV0130ReportNaming:
    def test_export_filenames_identifiable(self, client):
        import urllib.parse as up
        diff_id, _ = TestDiffFlow().run_diff(client)
        cases = [
            ("/api/export/report/" + diff_id, {"format": "csv"}, "照合レポート_master_vs_sf_"),
            ("/api/export/verify/" + diff_id, {"format": "xlsx"}, "投入検証_master_vs_sf_"),
            ("/api/export/html/" + diff_id, {}, "照合レポート_master_vs_sf_"),
            ("/api/export/upsert/" + diff_id, {"external_id": "社員番号"}, "アップサート_master_vs_sf_"),
        ]
        for path, body, prefix in cases:
            r = client.post(path, json=body)
            assert r.status_code == 200, r.text
            cd = up.unquote(r.headers["content-disposition"])
            assert prefix in cd, cd
            # タイムスタンプ(YYYYMMDD_HHMM)付き
            import re as _re
            assert _re.search(r"\d{8}_\d{4}\.(csv|xlsx|html)", cd), cd

    def test_rich_html_report(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.post(f"/api/export/html/{diff_id}", json={})
        text = r.content.decode("utf-8")
        # リッチ要素: 進捗バー・サマリーカード・列ランキング・旧→新・凡例・フィルタ
        for marker in ("progress", "cards", "差異の多い項目", "old", "要確認のみ",
                       "凡例", "レコード中", "DiffDesk v"):
            assert marker in text, marker
        # 変更セルの旧→新データが入っている(0002のメール)
        assert "hanako@example.com" in text and "hanako-new@example.com" in text


class TestV0131ValueRuleCount:
    def test_count_endpoint(self, client):
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.get(f"/api/diff/{diff_id}/value-rule-count",
                       params={"col_a": "メール", "value_a": "hanako@example.com",
                               "value_b": "hanako-new@example.com"})
        assert r.status_code == 200
        assert r.json()["count"] == 1
        r2 = client.get(f"/api/diff/{diff_id}/value-rule-count",
                        params={"col_a": "メール", "value_a": "x", "value_b": "y"})
        assert r2.json()["count"] == 0


class TestV0140KeyMode:
    def upload_pair(self, client):
        fa = upload(client, "a.csv",
                    "名前,値\n山田,1\n佐藤,2\n鈴木,3\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv",
                    "name,val\n山田,1\n佐藤,9\n".encode("utf-8"))["file_id"]
        return fa, fb

    def keyless_pairs(self):
        return [{"col_a": "名前", "col_b": "name"},
                {"col_a": "値", "col_b": "val"}]

    def test_row_number_mode(self, client):
        fa, fb = self.upload_pair(client)
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb,
            "mapping": {"pairs": self.keyless_pairs(), "key_mode": "row_number"},
        })
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["key_mode"] == "row_number"
        s = body["summary"]
        assert s["same"] == 1 and s["changed"] == 1 and s["only_a"] == 1
        rows = client.get(f"/api/diff/{body['diff_id']}/rows").json()["rows"]
        assert rows[0]["key"] == ["行1"]

    def test_content_mode(self, client):
        fa, fb = self.upload_pair(client)
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb,
            "mapping": {"pairs": self.keyless_pairs(), "key_mode": "content"},
        })
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["key_mode"] == "content"
        s = body["summary"]
        # 内容一致: 山田のみ一致、佐藤2/鈴木3はAのみ、佐藤9はBのみ
        assert s["same"] == 1 and s["changed"] == 0
        assert s["only_a"] == 2 and s["only_b"] == 1

    def test_columns_mode_keyless_rejected(self, client):
        fa, fb = self.upload_pair(client)
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb,
            "mapping": {"pairs": self.keyless_pairs(), "key_mode": "columns"},
        })
        assert r.status_code == 400
        assert "キー" in r.json()["error"]["message"]

    def test_default_key_mode_in_response(self, client):
        _, body = TestDiffFlow().run_diff(client)
        assert body["key_mode"] == "columns"


class TestV0160ProjectsUndo:
    def test_projects_crud(self, client):
        r = client.get("/api/projects").json()
        assert r["current"] == "既定"
        r = client.post("/api/projects", json={"name": "案件A"}).json()
        assert r["current"] == "案件A"
        r = client.post("/api/projects/switch", json={"name": "既定"}).json()
        assert r["current"] == "既定"
        r = client.delete("/api/projects/案件A").json()
        assert "案件A" not in r["names"]
        assert client.post("/api/projects", json={"name": ""}).status_code == 400
        assert client.delete("/api/projects/既定").status_code == 400

    def test_known_diffs_isolated_by_project(self, client):
        client.post("/api/known-diffs", json={"entry": {
            "type": "value", "col_a": "c", "value_a": "1", "value_b": "1.0"}})
        assert len(client.get("/api/known-diffs").json()["entries"]) == 1
        client.post("/api/projects", json={"name": "別案件"})
        assert client.get("/api/known-diffs").json()["entries"] == []
        client.post("/api/projects/switch", json={"name": "既定"})
        assert len(client.get("/api/known-diffs").json()["entries"]) == 1

    def test_undo_endpoint(self, client):
        assert client.get("/api/undo").json()["count"] == 0
        assert client.post("/api/undo").status_code == 400
        client.post("/api/known-diffs", json={"entry": {
            "type": "value", "col_a": "c", "value_a": "1", "value_b": "1.0"}})
        peek = client.get("/api/undo").json()
        assert peek["count"] == 1 and "既知差分" in peek["label"]
        r = client.post("/api/undo").json()
        assert "既知差分" in r["label"]
        assert client.get("/api/known-diffs").json()["entries"] == []

    def test_update_check(self, client, monkeypatch):
        import diffdesk.web.routes as routes_mod
        from diffdesk import __version__
        monkeypatch.setattr(routes_mod, "_fetch_latest_version", lambda: "99.0.0")
        monkeypatch.setitem(routes_mod._update_cache, "at", 0.0)
        r = client.get("/api/update-check").json()
        assert r["current"] == __version__
        assert r["latest"] == "99.0.0" and r["update_available"] is True
        # オフライン(取得失敗)でも壊れない
        def boom():
            raise OSError("offline")
        monkeypatch.setattr(routes_mod, "_fetch_latest_version", boom)
        monkeypatch.setitem(routes_mod._update_cache, "at", 0.0)
        r = client.get("/api/update-check").json()
        assert r["latest"] is None and r["update_available"] is False


class TestV0170Worksession:
    def test_save_restore_roundtrip(self, client):
        fa = upload(client, "a.csv", "id,v\n1,x\n2,y\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv", "id,v\n1,x\n2,z\n".encode("utf-8"))["file_id"]
        r = client.post("/api/worksession/save", json={
            "name": "作業1", "file_a": fa, "file_b": fb,
            "mapping": {"pairs": [{"col_a": "id", "col_b": "id", "is_key": True},
                                  {"col_a": "v", "col_b": "v"}],
                        "key_mode": "columns"},
            "options": {"trim": True}, "row_filter": {},
        })
        assert r.status_code == 200, r.text
        assert len(r.json()["files"]) == 2

        r = client.get("/api/worksessions").json()
        assert r["sessions"][0]["name"] == "作業1"

        r = client.post("/api/worksession/restore", json={"name": "作業1"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["role_a"] and body["role_b"]
        assert body["mapping"]["key_mode"] == "columns"
        # 復元されたファイルでそのまま差分が実行できる
        d = client.post("/api/diff", json={
            "file_a": body["role_a"], "file_b": body["role_b"],
            "mapping": body["mapping"],
        })
        assert d.status_code == 200
        assert d.json()["summary"]["changed"] == 1

        r = client.delete("/api/worksession/作業1")
        assert r.status_code == 200
        assert client.get("/api/worksessions").json()["sessions"] == []

    def test_restore_missing_404ish(self, client):
        r = client.post("/api/worksession/restore", json={"name": "ない"})
        assert r.status_code == 400

    def test_sessions_scoped_to_project(self, client):
        upload(client, "a.csv", "id\n1\n".encode("utf-8"))
        client.post("/api/worksession/save", json={"name": "s1"})
        client.post("/api/projects", json={"name": "別案件"})
        assert client.get("/api/worksessions").json()["sessions"] == []
        client.post("/api/projects/switch", json={"name": "既定"})
        assert len(client.get("/api/worksessions").json()["sessions"]) == 1


class TestV0180AuditPack:
    def test_audit_recorded_and_listed(self, client):
        TestDiffFlow().run_diff(client)
        client.post("/api/known-diffs", json={"entry": {
            "type": "value", "col_a": "メール", "value_a": "a", "value_b": "b"}})
        records = client.get("/api/audit").json()["records"]
        actions = [r["action"] for r in records]
        assert "照合実行" in actions and "既知差分登録" in actions
        r = client.post("/api/export/audit", json={})
        assert r.status_code == 200
        assert "監査ログ" in __import__("urllib.parse", fromlist=["unquote"]).unquote(
            r.headers["content-disposition"])

    def test_pack_export(self, client):
        import io
        import zipfile
        diff_id, _ = TestDiffFlow().run_diff(client)
        r = client.post(f"/api/export/pack/{diff_id}", json={})
        assert r.status_code == 200, r.text
        import urllib.parse as up
        cd = up.unquote(r.headers["content-disposition"])
        assert "検証パック_master_vs_sf_" in cd and cd.endswith('.zip')
        z = zipfile.ZipFile(io.BytesIO(r.content))
        assert "はじめにお読みください.txt" in z.namelist()
        assert "検証レポート.xlsx" in z.namelist()


class TestV0191HelpPage:
    def test_help_page_served(self, client):
        r = client.get("/help")
        assert r.status_code == 200
        text = r.text
        assert "DiffDesk ヘルプ" in text
        from diffdesk import __version__
        assert f"v{__version__}" in text  # {{V}}が置換されている
        # 主要セクションが揃っている
        for anchor in ("はじめに", "既知差分", "手動紐づけ", "検証パック",
                       "ショートカット", "よくある質問", "多対多検証"):
            assert anchor in text, anchor


class TestV0200ClassifySearch:
    def upload_diff(self, client):
        fa = upload(client, "a.csv",
                    "id,名前,数\n1,山田 太郎,100\n2,佐藤花子,200.0\n3,鈴木一郎,300\n"
                    .encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv",
                    "id,名前,数\n1,山田太郎,100\n2,佐藤花子,200\n3,田中次郎,300\n"
                    .encode("utf-8"))["file_id"]
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb,
            "mapping": {"pairs": [
                {"col_a": "id", "col_b": "id", "is_key": True},
                {"col_a": "名前", "col_b": "名前"},
                {"col_a": "数", "col_b": "数"}]},
            "options": {"trim": False, "normalize_width": False,
                        "normalize_numeric": False},
        })
        assert r.status_code == 200, r.text
        return r.json()["diff_id"]

    def test_classify_endpoint(self, client):
        diff_id = self.upload_diff(client)
        r = client.get(f"/api/diff/{diff_id}/classify").json()
        by = {c["cause"]: c["count"] for c in r["causes"]}
        assert by == {"space": 1, "numeric": 1, "real": 1}

    def test_known_by_cause_and_undo(self, client):
        diff_id = self.upload_diff(client)
        r = client.post(f"/api/diff/{diff_id}/known-by-cause",
                        json={"cause": "numeric"})
        assert r.json()["registered"] == 1
        # 分類から消える
        r = client.get(f"/api/diff/{diff_id}/classify").json()
        assert "numeric" not in {c["cause"] for c in r["causes"]}
        # realは拒否
        assert client.post(f"/api/diff/{diff_id}/known-by-cause",
                           json={"cause": "real"}).status_code == 400
        # アンドゥ1回で全て戻る
        client.post("/api/undo")
        r = client.get(f"/api/diff/{diff_id}/classify").json()
        assert "numeric" in {c["cause"] for c in r["causes"]}

    def test_rows_search_and_filters(self, client):
        diff_id = self.upload_diff(client)
        # 検索(値)
        r = client.get(f"/api/diff/{diff_id}/rows", params={"q": "花子"}).json()
        assert r["total"] == 1 and r["rows"][0]["key"] == ["2"]
        # 検索(キー)
        r = client.get(f"/api/diff/{diff_id}/rows", params={"q": "3"}).json()
        assert r["total"] == 1
        # 原因フィルタ
        r = client.get(f"/api/diff/{diff_id}/rows",
                       params={"cause": "space"}).json()
        assert r["total"] == 1 and r["rows"][0]["key"] == ["1"]
        # 列フィルタ
        r = client.get(f"/api/diff/{diff_id}/rows",
                       params={"col": "数", "status": "changed"}).json()
        assert r["total"] == 1 and r["rows"][0]["key"] == ["2"]


class TestV0210PrevCompare:
    def run_pair(self, client, content_b):
        fa = upload(client, "m.csv", "id,v\n1,x\n2,y\n3,z\n".encode("utf-8"))["file_id"]
        fb = upload(client, "s.csv", content_b.encode("utf-8"))["file_id"]
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb,
            "mapping": {"pairs": [{"col_a": "id", "col_b": "id", "is_key": True},
                                  {"col_a": "v", "col_b": "v"}]}})
        return r.json()["diff_id"]

    def test_compare_prev_flow(self, client):
        d1 = self.run_pair(client, "id,v\n1,x\n2,WRONG\n")
        r = client.get(f"/api/diff/{d1}/compare-prev").json()
        assert r["available"] is False  # 初回は前回なし
        d2 = self.run_pair(client, "id,v\n1,BROKEN\n2,y\n")
        r = client.get(f"/api/diff/{d2}/compare-prev").json()
        assert r["available"] and r["counts"] == {
            "new": 1, "resolved": 1, "continuing": 1}
        # 前回比較フィルタで行取得
        rows = client.get(f"/api/diff/{d2}/rows", params={"prev": "new"}).json()
        assert rows["total"] == 1 and rows["rows"][0]["key"] == ["1"]
        rows = client.get(f"/api/diff/{d2}/rows",
                          params={"prev": "continuing"}).json()
        assert rows["total"] == 1 and rows["rows"][0]["key"] == ["3"]

    def test_preflight_endpoint(self, client):
        fa = upload(client, "a.csv", "id,v\n1,x\n2,y\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv", "id,v\n1,x\n1,x\n".encode("utf-8"))["file_id"]
        r = client.post("/api/preflight", json={"file_a": fa, "file_b": fb})
        assert r.status_code == 200
        msgs = " / ".join(c["message"] for c in r.json()["checks"])
        assert "B側に重複" in msgs


class TestV0220RefPiiAi:
    def test_refcheck_and_export(self, client):
        fc = upload(client, "child.csv",
                    "id,dept\n1,D01\n2,D99\n3,\n".encode("utf-8"))["file_id"]
        fm = upload(client, "master.csv",
                    "code\nD01\nD02\n".encode("utf-8"))["file_id"]
        body = {"file_child": fc, "col_child": "dept",
                "file_master": fm, "col_master": "code"}
        r = client.post("/api/refcheck", json=body).json()
        assert r["missing"] == 1 and r["blank"] == 1 and r["matched"] == 1
        r = client.post("/api/export/refcheck", json=body)
        assert r.status_code == 200
        text = r.content.decode("utf-8-sig")
        assert "D99" in text and "エラー理由" in text

    def make_gender_diff(self, client):
        fa = upload(client, "a.csv",
                    "id,氏名,性別\n1,山田,男\n2,佐藤,女\n".encode("utf-8"))["file_id"]
        fb = upload(client, "b.csv",
                    "id,氏名,性別\n1,山田,Male\n2,佐藤,Female\n".encode("utf-8"))["file_id"]
        r = client.post("/api/diff", json={
            "file_a": fa, "file_b": fb,
            "mapping": {"pairs": [{"col_a": "id", "col_b": "id", "is_key": True},
                                  {"col_a": "氏名", "col_b": "氏名"},
                                  {"col_a": "性別", "col_b": "性別"}]}})
        return r.json()["diff_id"]

    def test_pii_endpoint(self, client):
        diff_id = self.make_gender_diff(client)
        r = client.get(f"/api/diff/{diff_id}/pii").json()
        assert {"column": "氏名", "kind": "氏名"} in r["columns"]

    def test_ai_sort_roundtrip(self, client):
        diff_id = self.make_gender_diff(client)
        r = client.get(f"/api/diff/{diff_id}/known-prompt").json()
        assert r["count"] == 2 and "性別" in r["prompt"]
        ans = '[{"no":1,"verdict":"known","reason":"変換"},{"no":2,"verdict":"known","reason":"変換"}]'
        r = client.post(f"/api/diff/{diff_id}/known-answer", json={"text": ans}).json()
        assert len(r["candidates"]) == 2
        r = client.post(f"/api/diff/{diff_id}/known-bulk",
                        json={"entries": r["candidates"] and [
                            {"col": c["col"], "value_a": c["value_a"],
                             "value_b": c["value_b"], "note": "AI承認"}
                            for c in r["candidates"]]}).json()
        assert r["registered"] == 2
        # 登録後は照合で既知扱い
        v = client.get(f"/api/diff/{diff_id}/verify").json()
        assert v["passed"] is True
        # アンドゥ1回で戻る
        client.post("/api/undo")
        assert client.get("/api/known-diffs").json()["entries"] == []
